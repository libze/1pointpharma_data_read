from pandas.api import types as ptypes
#from reader import cached_read_excel, get_input_rows
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
import re
import os
import numpy as np
import string
import math
from openpyxl import load_workbook
import random as pyrandom
import requests
# Temp numbers
MONTH_SOURCING = 3
PROCENT_UPPERBOUND = 400

# --- SEMANTIC_GROUPS ---
SEMANTIC_GROUPS = {
    "Brand name": ["Brand", "Brand name", "Brand name.1", "LocalName", "Handelsnavn"],
    "Active ingredient": ["Virkestoff", "Active ingredient", "Active ingredient.1", "ActiveIngredient"],
    "Supplier": ["Supplier", "supplier"],
    "Strength": ["Styrke", "Strength", "Strength.1", "StrengthText"],
    "Form": ["Legemiddelform", "Form", "Form.1"],

    # Pack Size no longer includes Pakningstype
    "Pack Size": ["Pack Size", "Pack Size.1", "PacksizeText"],


    "MA Holder": ["Innehaver", "MA Holder", "MA Holder.1"],
    "Country": ["Country", "Country pack", "Country Pack", "CountryPack"],
    "ATC code": ["ATC-kode (pakning)", "ATC code", "ATC Code", "ATCCode"],
    "Local item number / PZN number": [
        "Local item number / PZN number",
        "Local item number / PZN number.1"
    ],
    "Quantity Offered": ["Quantity Offered", "Quantity Offered.1"],
    "Quantatity requested": ["Quantatity requested"],
}

# --- PREFERRED_ORDER ---
PREFERRED_ORDER = {
    "Brand name": ["Brand name", "LocalName", "Handelsnavn", "Brand", "Brand name.1"],
    "Supplier": ["Supplier", "supplier"],
    "Active ingredient": ["Active ingredient", "ActiveIngredient", "Virkestoff", "Active ingredient.1"],
    "Strength": ["Strength", "StrengthText", "Styrke", "Strength.1"],
    "Form": ["Form", "Legemiddelform", "Form.1"],
    "Pack Size": ["Pack Size", "PacksizeText", "Pack Size.1"],
    "MA Holder": ["MA Holder", "Innehaver", "MA Holder.1"],
    "Country": ["Country", "Country pack", "Country Pack", "CountryPack"],
    "ATC code": ["ATC code", "ATCCode", "ATC Code", "ATC-kode (pakning)"],
    "Local item number / PZN number": [
        "Local item number / PZN number",
        "Local item number / PZN number.1"
    ],
    "Quantity Offered": ["Quantity Offered", "Quantity Offered.1"],
    "Quantatity requested": ["Quantatity requested"],
}

STRENGTH_COLS_BY_SOURCE = {
    "product_catalog": ["StrengthText"],
    "legemidler": ["Styrke"],
    "special_access": ["Strength"],         # if present in that DF branch
    "sourcing": ["Strength"],                         # none
    "topp 500": [],                         # none
}

DROP_ALWAYS = [
    "Multippel", "Trinnpris Gyldig", "Refusjonspris Gyldig", "Reseptgruppe",
    "Markedsføringsstatus", "Maks AUP Gyldig",
    "Local item number / PZN number",  # main label
    "target price EUR",
    "Is it a blood product?", "Is it a controlled drug?", "Is it registered within EU?",
    "Documents needed?", "Comments", "Customer name", "Customer country",
    "Shelf Life", "Expiry date", "Storage Requirements", "Supplier comments",
    "Quoted Customer?", "Reason (if no)"
]


# alternate spellings we'll accept for the quantity column
QTY_LAST_12M_CANDIDATES = [
    "Qty. Sold last 12 months - Retail",
    "Qty Sold last 12 months - Retail",
    "qty. Sold last 12 months - Retail",
    "Units sold last 12 months - Retail",
]

# --- Live FX (fallbacks used if HTTP fails) ---
DEFAULT_RATES = {"EUR_DKK": 7.46, "NOK_DKK": 0.66}
_RATES_CACHE = None

def fetch_exchange_rates() -> dict:
    """
    Fetch DKK rates live:
      - EUR -> DKK
      - NOK -> DKK
    Falls back to DEFAULT_RATES if anything fails.
    """
    try:
        r1 = requests.get("https://api.frankfurter.app/latest",
                          params={"from": "EUR", "to": "DKK"}, timeout=8)
        r1.raise_for_status()
        eur_dkk = float(r1.json()["rates"]["DKK"])

        r2 = requests.get("https://api.frankfurter.app/latest",
                          params={"from": "NOK", "to": "DKK"}, timeout=8)
        r2.raise_for_status()
        nok_dkk = float(r2.json()["rates"]["DKK"])

        return {"EUR_DKK": eur_dkk, "NOK_DKK": nok_dkk}
    except Exception:
        return DEFAULT_RATES

def get_rates() -> dict:
    global _RATES_CACHE
    if _RATES_CACHE is None:
        _RATES_CACHE = fetch_exchange_rates()
    return _RATES_CACHE

_NUM_RE = re.compile(r"^\s*[+-]?(\d{1,3}([ ,]\d{3})*|\d+)([.,]\d+)?\s*$")

def extract_requested_quantity(q):
    """
    Extracts the first integer-like value from a messy quantity string.
    Handles dots, commas, and trailing text.
    """
    if pd.isna(q):
        return None

    s = str(q).lower().replace(".", "").replace(" ", "")  # remove . and spaces
    s = re.sub(r"[^\d,]", "", s)  # keep digits and commas only
    s = s.replace(",", "")  # remove commas (treat all as thousands sep)

    try:
        return int(s)
    except ValueError:
        return None


def _drop_unwanted_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    cols = [c for c in DROP_ALWAYS if c in df.columns]
    return df.drop(columns=cols, errors="ignore")
def _to_num(x):
    ok, v = _try_parse_number(x)
    return v if ok else np.nan

def _keep_best_strength_rows(df: pd.DataFrame, input_strength: str, source_name: str) -> pd.DataFrame:
    """
    Use only the first available strength column for this source and
    keep rows with the highest score; if all scores are 0, keep all.
    """

    cols = [c for c in STRENGTH_COLS_BY_SOURCE.get(source_name, []) if c in df.columns]
    if df.empty or not input_strength or not cols:
        return df

    sc = cols[0]
    scores = df[sc].astype(str).map(lambda s: _strength_match_score(input_strength, s))
    max_score = scores.max() if len(scores) else 0
    return df if (pd.isna(max_score) or max_score <= 0) else df.loc[scores == max_score]
def _unique_sheet_name(base, path):
    """Return a sheet name that is unique within the workbook."""


    # Excel sheet names are limited to 31 chars
    base = base[:25]  # leave room for suffix
    name = base
    if os.path.exists(path):
        try:
            wb = load_workbook(path, read_only=True)
            existing = set(wb.sheetnames)
            wb.close()
        except Exception:
            existing = set()
        while name in existing:
            suffix = "_" + "".join(pyrandom.choices(string.ascii_uppercase + string.digits, k=4))
            name = (base + suffix)[:31]
    return name
def parse_strength(s):
    """
    Parses a strength string into list of (value, unit) tokens.
    e.g. "500mg/5ml" -> [(500.0, "mg"), (5.0, "ml")]
    """
    s = str(s).lower().replace(",", ".")
    tokens = re.findall(r'(\d+(?:\.\d+)?)\s*([a-zA-Zμ/]+)', s)
    parsed = []
    for val, unit in tokens:
        try:
            parsed.append((float(val), unit.strip()))
        except ValueError:
            print('An error has occurred when looking at strength units. Continuing')
            continue
    return parsed

def strength_tokens_match(input_str, catalog_str):
    parsed_input = parse_strength(input_str)
    parsed_catalog = parse_strength(catalog_str)

    if parsed_input == parsed_catalog:
        return True

    # Optional: implement simplified ratio matching
    if len(parsed_input) == len(parsed_catalog) == 2:
        # Match e.g. 500mg/5ml == 100mg/1ml
        (v1, u1), (v2, u2) = parsed_input
        (v3, u3), (v4, u4) = parsed_catalog
        if u1 == u3 and u2 == u4:
            return abs((v1/v2) - (v3/v4)) < 0.01  # tolerance for float rounding
    return False


# Sourcing look up logic
def sourcing(input_rows, sourcing_data):
    sourcing_data = sourcing_data.copy()

    # Normalize sourcing_data
    sourcing_data['Brand name'] = sourcing_data['Brand name'].astype(str).str.strip().str.lower()
    sourcing_data['Quote Date'] = pd.to_datetime(sourcing_data['Quote Date'], errors='coerce')
    sourcing_data['Quantity Offered'] = pd.to_numeric(sourcing_data['Quantity Offered'], errors='coerce')

    # Remove last N rows
    source_no_input = sourcing_data[:-len(input_rows)]

    """"
    Dont need to cutoff n months
    
    # Only include quotes within the last N months
    cutoff_date = datetime.today() - relativedelta(months=MONTH_SOURCING)
    source_no_input = source_no_input[source_no_input['Quote Date'] >= cutoff_date]
    """

    results = {}

    for _, row in input_rows.iterrows():
        original_brand = row['Brand name']
        #brand = str(original_brand).strip().lower()
        brand = str(original_brand).strip().lower().split()[0]
        requested_q = extract_requested_quantity(row['Quantatity requested'])

        # Match rows by brand
        matches = source_no_input[
            source_no_input['Brand name']
            .astype(str)
            .str.lower()
            .str.contains(brand, na=False)
        ].copy()

        if matches.empty:
            results[original_brand] = pd.DataFrame()
            continue

        # Apply quantity check if available
        if requested_q is not None:
            allowed = (
                    matches['Quantity Offered'].notna() &
                    (requested_q <= matches['Quantity Offered'] * (1 + PROCENT_UPPERBOUND / 100))
            )
            matches = matches[allowed]

        results[brand] = matches

    return results

def product_catalog(input_rows, product_catalog_df):
    product_catalog_df = product_catalog_df.copy()
    product_catalog_df['LocalName'] = product_catalog_df['LocalName'].astype(str).str.strip().str.lower()
    product_catalog_df['ActiveIngredient'] = product_catalog_df['ActiveIngredient'].astype(str).str.strip().str.lower()
    product_catalog_df['StrengthText'] = product_catalog_df['StrengthText'].astype(str).str.strip().str.lower()

    matches = {}

    for idx, row in input_rows.iterrows():
        brand_raw = str(row['Brand name']).strip().lower()
        brand_first_word = brand_raw.split()[0] if brand_raw else ""
        ingredient = str(row['Active ingredient']).strip().lower()
        strength = str(row['Strength']).strip().lower()

        # Step 1: match first word of brand in LocalName
        match = product_catalog_df[
            product_catalog_df['LocalName'].str.contains(brand_first_word, case=False, regex=False, na=False)
        ]

        # Step 2: fallback to active ingredient
        if match.empty:
            match = product_catalog_df[
                product_catalog_df['ActiveIngredient'] == ingredient
            ]

        # Step 3: narrow by strength
        if len(match) > 1:
            original_match = match.copy()
            match = match[
                match['StrengthText'].apply(lambda s: strength_tokens_match(strength, s))
            ]
            if match.empty:
                match = original_match

        matches[brand_raw] = match

    return matches


def legemidler(input_rows, legemidler_df):
    legemidler_df = legemidler_df.copy()
    legemidler_df['Handelsnavn'] = legemidler_df['Handelsnavn'].astype(str).str.strip().str.lower()
    legemidler_df['Virkestoff'] = legemidler_df['Virkestoff'].astype(str).str.strip().str.lower()
    legemidler_df['Styrke'] = legemidler_df['Styrke'].astype(str).str.strip().str.lower()

    matches = {}

    for idx, row in input_rows.iterrows():
        brand_raw = str(row['Brand name']).strip().lower()
        brand = brand_raw.split()[0] if brand_raw else ""
        ingredient = str(row['Active ingredient']).strip().lower()
        strength = str(row['Strength']).strip().lower()

        # Step 1: match first word of brand in Handelsnavn
        match = legemidler_df[
            legemidler_df['Handelsnavn'].str.contains(brand, case=False, regex=False, na=False)
        ]

        # Step 2: fallback to exact match on Virkestoff
        if match.empty:
            match = legemidler_df[
                legemidler_df['Virkestoff'] == ingredient
            ]

        # Step 3: filter by strength
        if len(match) > 1:
            original_match = match.copy()
            match = match[
                match['Styrke'].apply(lambda s: strength_tokens_match(strength, s))
            ]
            if match.empty:
                match = original_match

        matches[brand] = match

    return matches

def check_availability_no(input_rows, availability_df):
    availability_df = availability_df.copy()
    availability_df['vare'] = availability_df['vare'].astype(str).str.strip().str.lower()

    results = {}

    for _, row in input_rows.iterrows():
        brand_raw = row['Brand name']
        brand_first_word = str(brand_raw).strip().lower().split()[0]
        requested_qty = extract_requested_quantity(row['Quantatity requested'])

        # Find match in availability table using first word
        match = availability_df[availability_df['vare'].str.contains(brand_first_word, na=False)]

        if match.empty:
            results[brand_first_word] = {
                'available': False,
                'reason': 'Brand not found',
                'price': None
            }
            continue

        match_row = match.iloc[0]
        mnd_av = str(match_row['Mnd AV']).strip()
        price = match_row['Pris NOK']

        # Default outcome
        outcome = {
            'available': True,
            'reason': 'Within limit or no limit',
            'price': price
        }

        if '+' in mnd_av:
            outcome['reason'] = 'Unlimited (+)'
        else:
            try:
                upper = int(mnd_av.replace(" ", "").split('-')[1])
                if requested_qty is not None:
                    if requested_qty <= upper:
                        outcome['reason'] = f'Within range (≤ {upper})'
                    else:
                        outcome['available'] = False
                        outcome['reason'] = f'Requested {requested_qty} > {upper}'
                else:
                    outcome['reason'] = 'Brand found, but could not parse quantity'
            except (IndexError, ValueError):
                outcome['reason'] = 'Could not parse Mnd AV'

        results[brand_first_word] = outcome

    return results

def special_access(input_rows, special_access_df):
    """
    Return a dict[brand] -> DataFrame of ALLOWED rows only.
    Rows include normal fields (Brand name, Active ingredient, Strength,
    Pack Size/PacksizeText, Form, Country, Supplier, etc.) so the rest of the
    pipeline (price calc, harmonization, strength filter) can treat it like other sources.
    """
    df = special_access_df.copy()

    # Normalize fields used for matching
    df['Brand name']        = df['Brand name'].astype(str).str.strip().str.lower()
    df['Active ingredient'] = df['Active ingredient'].astype(str).str.strip().str.lower()
    df['Strength']          = df['Strength'].astype(str).str.strip()

    # Keep pricing columns numeric-friendly for later price calc
    if 'target price EUR' in df.columns:
        df['target price EUR'] = (
            df['target price EUR'].astype(str)
            .str.replace(",", "", regex=False)
        )

    results = {}

    for _, row in input_rows.iterrows():
        brand_raw = str(row['Brand name']).strip().lower()
        brand_key = brand_raw.split()[0] if brand_raw else ""
        ingredient = str(row['Active ingredient']).strip().lower()
        input_strength = str(row['Strength']).strip()

        # Step 1: brand match by first word in "Brand name"
        match = df[df['Brand name'].str.contains(brand_key, case=False, regex=False, na=False)]

        # Step 2: fallback to ingredient if brand missing
        if match.empty:
            match = df[df['Active ingredient'] == ingredient]

        if match.empty:
            results[brand_key] = pd.DataFrame()
            continue

        # Step 3: ALLOWED filter based on requested quantity vs the sheet’s max
        requested_qty = extract_requested_quantity(row.get('Quantatity requested'))
        match = match.copy()
        if 'Quantatity requested' in match.columns:
            max_q = match['Quantatity requested'].apply(extract_requested_quantity)
            # Allow when either side can't be parsed OR requested <= max
            allowed_mask = max_q.isna() | (pd.isna(requested_qty)) | (requested_qty <= max_q)
            match = match.loc[allowed_mask]

        # If nothing allowed, keep empty
        if match.empty:
            results[brand_key] = pd.DataFrame()
            continue

        # Optional: keep a lowercase 'supplier' column for convenience
        if 'Supplier' in match.columns and 'supplier' not in match.columns:
            match['supplier'] = match['Supplier']

        results[brand_key] = match

    return results


def present_matches(all_matches: dict, label: str):
    print(f"\n--- {label} ---")
    found = False
    for key, df in all_matches.items():
        if isinstance(df, pd.DataFrame) and not df.empty:
            print(f"\n🔹 Match for: {key}")
            print(df)  # If using Jupyter, else use print(df)
            found = True
        elif isinstance(df, list) and df:  # for list-of-matches (like special access)
            print(f"\n🔹 Match for: {key}")
            for match in df:
                print(match)
            found = True
        elif isinstance(df, dict) and df.get("available"):  # for special_access dict format
            print(f"\n🔹 Match for: {key}")
            print(df)
            found = True
    if not found:
        print("No matches found.")

def present_grouped_matches(all_matches):
    # Collect all brands from all sources
    all_brands = set()
    for source_matches in all_matches.values():
        all_brands.update(source_matches.keys())

    for brand in sorted(all_brands):
        print(f"\n=== 🔍 Matches for: {brand} ===")
        brand_found = False

        for source_name, source_matches in all_matches.items():
            match = source_matches.get(brand)
            if match is None:
                continue

            print(f"\n--- Source: {source_name} ---")
            if isinstance(match, pd.DataFrame) and not match.empty:
                print(match)
                brand_found = True
            elif isinstance(match, list) and match:
                for m in match:
                    print(m)
                brand_found = True
            elif isinstance(match, dict):
                # Special access style
                if match.get("available") and match.get("matches"):
                    for m in match["matches"]:
                        print(m)
                    brand_found = True
                elif match.get("matches"):
                    print(f"(Not available, but matches found):")
                    for m in match["matches"]:
                        print(m)
                    brand_found = True
        if not brand_found:
            print("No matches in any source.")

def _try_parse_number(x):
    """
    Try to parse a scalar as a float in a robust way.
    Accepts ints/floats directly; for strings:
    - allows thousands sep (space or comma)
    - allows decimal sep . or ,
    - ignores surrounding spaces/currency if the whole string looks numeric
    Returns (is_numeric: bool, value: float|None)
    """
    if pd.isna(x):
        return False, None
    if isinstance(x, (int, float, np.integer, np.floating)) and not isinstance(x, bool):
        # Normalize -0.0 etc.
        val = float(x)
        return True, val

    s = str(x).strip()
    # quick accept if it "looks numeric" as a whole token
    if _NUM_RE.match(s):
        # remove spaces and thousands commas, unify decimal comma to dot
        s2 = s.replace(" ", "").replace(",", "")
        try:
            return True, float(s2)
        except ValueError:
            pass

    # Last-resort: strip obvious currency/symbols at ends
    s3 = s.strip("€$krKR")
    if _NUM_RE.match(s3):
        s3 = s3.replace(" ", "").replace(",", "")
        try:
            return True, float(s3)
        except ValueError:
            pass

    return False, None

def _normalize_text(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if not s:
        return None
    return " ".join(s.split()).casefold()

def _first_non_null_by_order(df, order):
    present = [c for c in order if c in df.columns]
    if not present:
        return pd.Series([pd.NA] * len(df), index=df.index)

    # Iterate row-wise: pick the first non-null in the preferred order
    result = []
    for i in df.index:
        val = pd.NA
        for col in present:
            if not pd.isna(df.at[i, col]):
                val = df.at[i, col]
                break
        result.append(val)

    return pd.Series(result, index=df.index).infer_objects()

def harmonize_semantic_columns_numeric_aware(df: pd.DataFrame, preserve_variants: dict[str, set[str]] | None = None):
    preserve_variants = preserve_variants or {}
    df = df.copy()
    groups_with_conflict = set()
    conflict_cols = []

    for canon, variants in SEMANTIC_GROUPS.items():
        present = [c for c in variants if c in df.columns]
        if not present:
            continue

        order = [c for c in PREFERRED_ORDER.get(canon, []) if c in present] or present
        df[canon] = _first_non_null_by_order(df, order)

        conflict_mask = []
        conflict_detail = []
        for i in df.index:
            vals = [(col, df.at[i, col]) for col in present if not pd.isna(df.at[i, col])]
            if len(vals) <= 1:
                conflict_mask.append(False)
                conflict_detail.append({})
                continue

            numeric_flags, numeric_vals = [], []
            for _, v in vals:
                is_num, num = _try_parse_number(v)
                numeric_flags.append(is_num); numeric_vals.append(num)

            if all(numeric_flags):
                rounded = [round(n if n is not None else math.nan, 10) for n in numeric_vals]
                unique_nums = set(rounded)
                if len(unique_nums) > 1:
                    conflict_mask.append(True); conflict_detail.append({col: str(v) for col, v in vals})
                else:
                    conflict_mask.append(False); conflict_detail.append({})
            else:
                uniq = { _normalize_text(v) for _, v in vals if _normalize_text(v) is not None }
                if len(uniq) > 1:
                    conflict_mask.append(True); conflict_detail.append({col: str(v) for col, v in vals})
                else:
                    conflict_mask.append(False); conflict_detail.append({})

        conflict_mask = pd.Series(conflict_mask, index=df.index)

        if conflict_mask.any():
            groups_with_conflict.add(canon)
            colname = f"{canon}_conflict"
            df[colname] = conflict_detail
            conflict_cols.append(colname)
        else:
            # drop non-canonical variants, EXCEPT those explicitly preserved
            preserve = preserve_variants.get(canon, set())
            drop_cols = [c for c in present if (c != canon and c not in preserve)]
            if drop_cols:
                df.drop(columns=drop_cols, inplace=True, errors="ignore")

    return df, sorted(groups_with_conflict), sorted(conflict_cols)

# Sample structure to simulate the function update
def filter_columns_by_source(df, source_name):
    col_indices = {
        "product_catalog": [0,1,2,3,4,5,6,8,11],
        "legemidler":      [2,3,4,5,6,7,9,10,11,12,21],
    }

    always_keep = [
        "Supplier", "supplier",
        "Innehaver", "MA Holder", "Manufacturer",
        "Wholesale Purchase price", "Cost Price €", "target price EUR",
        "Maks AIP Gyldig", "Pris NOK", "Quantity Offered",
        "ATC code", "ATC Code", "ATCCode", "ATC-kode (pakning)",
        "Pack Size", "Pakningstype", "Country",
        "Quantatity requested",
        "Ref",
        # keep these if present:
        "Price EUR",                    # <-- ADD THIS (was missing)
        # (and if you truly don't want DKK anymore, you can drop the next two)
        # "Wholesale Purchase price (DKK)", "Unit price (EUR)", "Unit price (DKK)",
        # "WPP (DKK)",

        "Qty. Sold last 12 months - Retail",
        "Qty Sold last 12 months - Retail",
        "qty. Sold last 12 months - Retail",
        "Units sold last 12 months - Retail",
    ]

    if source_name not in col_indices:
        return df

    cols = df.columns.tolist()
    valid_idx = [i for i in col_indices[source_name] if i < len(cols)]
    sel = df.iloc[:, valid_idx].copy()
    for c in always_keep:
        if c in df.columns and c not in sel.columns:
            sel[c] = df[c]
    return sel



def apply_source_specific_enrichment(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if source_name != "legemidler":
        return df

    out = df.copy()
    out["Country"] = "NO"
    out = build_legemidler_packsize(out)

    manu = pd.Series(pd.NA, index=out.index, dtype="object")
    if "Innehaver" in out.columns:
        s = out["Innehaver"].astype(str).str.strip()
        manu = s.where(s.ne(""), pd.NA)

    if "MA Holder" in out.columns:
        s2 = out["MA Holder"].astype(str).str.strip()
        manu = manu.fillna(s2.where(s2.ne(""), pd.NA))

    out["Manufacturer"] = manu
    return out


def collapse_semantic_variants(
    df: pd.DataFrame,
    preserve_variants: dict[str, set[str]] | None = None,
) -> pd.DataFrame:
    """
    For each semantic group:
      - Ensure a single canonical column exists (rename first present variant if needed).
      - Drop other variants, except those listed in `preserve_variants[canonical]`.
        NOTE: preserve only applies when the canonical column exists; if we had to
        create it via rename, we drop all remaining variants to avoid duplication.
      - Finally, rename 'Brand name' -> 'Brand' only if 'Brand' doesn't already exist.
    """
    preserve_variants = preserve_variants or {}
    df = df.copy()

    for canon, variants in SEMANTIC_GROUPS.items():
        present = [v for v in variants if v in df.columns]

        # If neither canonical nor any variants are present, nothing to do.
        if not present and canon not in df.columns:
            continue

        # Ensure canonical column exists (rename one variant to canonical if needed).
        created_by_rename = False
        if canon not in df.columns and present:
            chosen = present[0]
            if chosen != canon:
                df.rename(columns={chosen: canon}, inplace=True)
                created_by_rename = True

        # Decide which variants to drop.
        if created_by_rename:
            # When we just created the canonical, do NOT preserve other variants
            # to avoid duplicated semantics.
            preserve = set()
        else:
            preserve = preserve_variants.get(canon, set())

        to_drop = [
            v for v in variants
            if v in df.columns and v != canon and v not in preserve
        ]
        if to_drop:
            df.drop(columns=to_drop, inplace=True, errors="ignore")

    # Rename 'Brand name' to 'Brand' if 'Brand' not already present.
    if "Brand name" in df.columns and "Brand" not in df.columns:
        df.rename(columns={"Brand name": "Brand"}, inplace=True)

    return df


def _pick_from_input(current_brand: str, cand_lower: list[tuple[str, str]]) -> str | None:
    """
    Return the input brand (original casing) if any candidate substring is found
    inside current_brand. Prefers longer matches first.
    """
    if not isinstance(current_brand, str) or not current_brand.strip():
        return None
    cb = current_brand.lower()
    for orig, low in cand_lower:
        if low and low in cb:
            return orig
    return None

def _get_qty_col(df: pd.DataFrame) -> str | None:
    return next((c for c in QTY_LAST_12M_CANDIDATES if c in df.columns), None)


def _add_price_columns(df: pd.DataFrame, source_name: str, rates) -> pd.DataFrame:
    """
    Compute a unified 'Price EUR' column from the various source price columns.
    Priority:
      1) 'Wholesale Purchase price' (EUR)
      2) 'Cost Price €'
      3) 'target price EUR'
      4) 'Maks AIP Gyldig' (NOK → EUR)
      5) 'Pris NOK' (NOK → EUR)
    """
    if df.empty:
        return df

    EUR_TO_DKK = rates["EUR_DKK"]
    NOK_TO_DKK = rates["NOK_DKK"]
    NOK_TO_EUR = NOK_TO_DKK / EUR_TO_DKK

    out = df.copy()

    wpp_eur  = out.get("Wholesale Purchase price", np.nan)
    cost_eur = out.get("Cost Price €", np.nan)
    targ_eur = out.get("target price EUR", np.nan)
    maks_aip = out.get("Maks AIP Gyldig", np.nan)
    pris_nok = out.get("Pris NOK", np.nan)
    qty_off  = out.get("Quantity Offered", np.nan)

    # convert to numeric
    wpp_eur  = pd.to_numeric(wpp_eur.apply(_to_num)  if isinstance(wpp_eur,  pd.Series) else pd.Series([], dtype=float), errors="coerce")
    cost_eur = pd.to_numeric(cost_eur.apply(_to_num) if isinstance(cost_eur, pd.Series) else pd.Series([], dtype=float), errors="coerce")
    targ_eur = pd.to_numeric(targ_eur.apply(_to_num) if isinstance(targ_eur, pd.Series) else pd.Series([], dtype=float), errors="coerce")
    maks_aip = pd.to_numeric(maks_aip.apply(_to_num) if isinstance(maks_aip, pd.Series) else pd.Series([], dtype=float), errors="coerce")
    pris_nok = pd.to_numeric(pris_nok.apply(_to_num) if isinstance(pris_nok, pd.Series) else pd.Series([], dtype=float), errors="coerce")
    qty_off  = pd.to_numeric(qty_off.apply(_to_num)  if isinstance(qty_off,  pd.Series) else pd.Series([], dtype=float), errors="coerce")

    # unified EUR price
    price_eur = pd.Series(np.nan, index=out.index, dtype=float)

    if len(wpp_eur):
        price_eur = np.where(~wpp_eur.isna(), wpp_eur, price_eur)

    if len(cost_eur):
        price_eur = np.where(np.isnan(price_eur) & ~cost_eur.isna(), cost_eur, price_eur)

    if len(targ_eur):
        price_eur = np.where(np.isnan(price_eur) & ~targ_eur.isna(), targ_eur, price_eur)

    if len(maks_aip):
        price_eur = np.where(np.isnan(price_eur) & ~maks_aip.isna(), maks_aip * NOK_TO_EUR, price_eur)

    if len(pris_nok):
        price_eur = np.where(np.isnan(price_eur) & ~pris_nok.isna(), pris_nok * NOK_TO_EUR, price_eur)

    out["Price EUR"] = pd.to_numeric(price_eur, errors="coerce").round(2)

    # Drop any old DKK/unit price columns
    to_drop = [c for c in ["WPP (DKK)", "Unit price (DKK)", "Unit price (EUR)"] if c in out.columns]
    if to_drop:
        out.drop(columns=to_drop, inplace=True, errors="ignore")

    return out


def _strength_match_score(input_strength: str, candidate_strength: str) -> int:
    """
    Score by longest common *prefix* length (characters) after normalization.
    Example: "250 iu + 5 ml" vs "250 iu" -> 6
    """
    if not input_strength or not candidate_strength:
        return 0
    a = _norm_strength_text(input_strength)
    b = _norm_strength_text(candidate_strength)
    return _lcp_len(a, b)

# Map each input brand's first token -> its Strength (as text)

def _build_strength_lookup(rows: pd.DataFrame) -> dict[str, str]:
    """
    Map <first word of Brand name> -> Strength from input rows.
    """
    lut = {}
    if "Brand name" in rows.columns and "Strength" in rows.columns:
        for _, r in rows.iterrows():
            b = str(r.get("Brand name", "")).strip().lower().split()[0]
            if b:
                lut[b] = str(r.get("Strength", "")).strip()
    return lut

def _postprocess_legemidler(df: pd.DataFrame) -> pd.DataFrame:
    """Country, Pack Size, Manufacturer(from Innehaver/MA Holder), ATC code."""
    if df is None or df.empty:
        return df
    out = df.copy()

    # Country + Pack Size first (needs Antall/Mengde/Måle-enhet/Pakningstype)
    out["Country"] = "NO"
    out = build_legemidler_packsize(out)

    # Manufacturer: prefer Innehaver, fallback MA Holder
    manu = pd.Series(pd.NA, index=out.index, dtype="object")
    if "Innehaver" in out.columns:
        s = out["Innehaver"].astype(str).str.strip()
        manu = s.where(s.ne(""), pd.NA)
    if "MA Holder" in out.columns:
        s2 = out["MA Holder"].astype(str).str.strip()
        manu = manu.fillna(s2.where(s2.ne(""), pd.NA))
    out["Manufacturer"] = manu

    # Canonical ATC column
    if "ATC code" not in out.columns:
        for c in ("ATC code", "ATC Code", "ATCCode", "ATC-kode (pakning)"):
            if c in out.columns:
                out["ATC code"] = out[c]
                break

    return out

def _norm_strength_text(s: str) -> str:
    # lowercase, collapse internal whitespace to single spaces
    return " ".join(str(s).strip().lower().split())

def _lcp_len(a: str, b: str) -> int:
    """Longest common prefix length (character-wise)."""
    n = min(len(a), len(b))
    i = 0
    while i < n and a[i] == b[i]:
        i += 1
    return i

def normalize_brand_using_input(final_df: pd.DataFrame, input_rows: pd.DataFrame) -> pd.DataFrame:
    """
    Overwrite 'Brand' in final_df with the short 'Brand name' from input_rows:
      1) By Ref (exact map) when possible.
      2) Fallback: substring match of input brand within final_df['Brand'] (case-insensitive).
    """
    df = final_df.copy()

    # Ensure Brand column exists
    if "Brand" not in df.columns:
        if "Brand name" in df.columns:
            df.rename(columns={"Brand name": "Brand"}, inplace=True)
        else:
            df["Brand"] = pd.NA

    # 1) Map by Ref
    if "Ref" in df.columns and "Ref" in input_rows.columns and "Brand name" in input_rows.columns:
        ref_map = input_rows.set_index("Ref")["Brand name"]
        mapped = df["Ref"].map(ref_map)
        df["Brand"] = mapped.combine_first(df["Brand"])

    # 2) Fallback: substring match
    if "Brand name" in input_rows.columns:
        candidates = (
            input_rows["Brand name"].astype(str).str.strip().dropna().unique().tolist()
        )
        # sort longest-first so “nulojix 250 mg …” doesn’t swallow “nulojix”
        candidates.sort(key=lambda s: len(s), reverse=True)
        cand_lower = [(c, c.lower()) for c in candidates]

        overwrite = df["Brand"].apply(lambda b: _pick_from_input(b, cand_lower))
        df["Brand"] = overwrite.combine_first(df["Brand"])

    return df


# Format numbers and strings
def _fmt(x):
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        return str(int(x)) if float(x).is_integer() else str(x)
    return "" if x is None else str(x)

def build_legemidler_packsize(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build Pack Size for legemidler source as:
        - if Antall empty or '1':   "Mengde Måle-enhet[, Pakningstype]"
        - else:                     "Antall x Mengde Måle-enhet[, Pakningstype]"
    Cleans NaNs and formats numbers nicely. Drops the qty source cols afterwards
    but keeps 'Pakningstype' column (useful to see separately).
    """
    needed_qty = ["Antall beholdere", "Mengde per beholder", "Måle-enhet"]
    if not any(c in df.columns for c in (needed_qty + ["Pakningstype"])):
        return df

    antall = df.get("Antall beholdere", pd.Series([""] * len(df))).map(_fmt).replace(r"^(nan|none|null|\s*)$", "", regex=True)
    mengde = df.get("Mengde per beholder", pd.Series([""] * len(df))).map(_fmt).replace(r"^(nan|none|null|\s*)$", "", regex=True)
    enhet  = df.get("Måle-enhet",          pd.Series([""] * len(df))).map(_fmt).replace(r"^(nan|none|null|\s*)$", "", regex=True)
    pakn   = df.get("Pakningstype",        pd.Series([""] * len(df))).map(_fmt).replace(r"^(nan|none|null|\s*)$", "", regex=True)

    out = []
    for a, m, e, p in zip(antall, mengde, enhet, pakn):
        # quantity part
        if a and a != "1":
            qty = f"{a} x {m} {e}".strip()
        else:
            qty = f"{m} {e}".strip()

        qty = " ".join(qty.split()).strip()  # normalize spaces

        # append pakningstype if present
        if p:
            combo = f"{qty}, {p}" if qty else p
        else:
            combo = qty

        out.append(combo if combo else pd.NA)

    df["Pack Size"] = pd.Series(out, index=df.index).replace(r"^\s*$", pd.NA, regex=True)

    # Drop the qty source cols (keep Pakningstype column)
    df.drop(columns=[c for c in (needed_qty + ["Pakningstype"]) if c in df.columns], inplace=True, errors="ignore")

    return df


def export_matches_to_excel(all_matches, rows, sheet_name=None, filename="match_results.xlsx", add_spaces=True):
    output_dir = 'Output'
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)
    rates = get_rates()

    if not sheet_name:
        base = datetime.now().strftime("Run_%Y-%m-%d_%H%M")
        sheet_name = _unique_sheet_name(base, output_path)

    strength_lookup = _build_strength_lookup(rows)

    # --- Build maps from input rows (use composite key to avoid clashes) ---
    _rows = rows.copy()

    # brand-first, case-insensitive
    _rows["_brand_first"] = (
        _rows["Brand name"].astype(str).str.strip().str.split().str[0].str.lower()
    )

    # normalized strength (uses your existing _norm_strength_text)
    _rows["_strength_norm"] = _rows["Strength"].astype(str).map(_norm_strength_text)

    # composite key: brand + strength
    _rows["_key"] = _rows["_brand_first"] + "||" + _rows["_strength_norm"]

    # raw requested qty string (keeps "10 (MOQ)")
    _rows["_req_str"] = _rows["Quantatity requested"].astype(str)

    # prefer the last non-empty per key
    nonempty = _rows[~_rows["_req_str"].str.strip().eq("")]
    qty_map_by_key = (
        nonempty.dropna(subset=["_key"]).groupby("_key", sort=False)["_req_str"].last()
    )

    # brand-only fallback (when no strength match)
    qty_map_by_brand = (
        nonempty.dropna(subset=["_brand_first"])
        .groupby("_brand_first", sort=False)["_req_str"]
        .last()
    )

    # New section
    all_rows = []

    for source_name, source_matches in all_matches.items():
        for brand, match in source_matches.items():
            if brand == 'nulojix':
                pass

            if isinstance(match, pd.DataFrame) and not match.empty:
                df = match.copy()

                # 1) compute prices BEFORE filtering drops raw cols
                df = _add_price_columns(df, source_name, rates)

                # 2) pack size BEFORE filtering (needs Antall/Mengde/Måle-enhet/Pakningstype)
                if source_name == "legemidler":
                    df = _postprocess_legemidler(df)

                # 3) keep the best strength rows (uses strength cols still present)
                brand_first = str(brand).strip().lower().split()[0] if isinstance(brand, str) else ""
                inp_strength = strength_lookup.get(brand_first, "")
                df = _keep_best_strength_rows(df, inp_strength, source_name)

                # 4) now trim columns (but preserve computed prices & strength via ALWAYS_KEEP)
                df = filter_columns_by_source(df, source_name)
                df = _drop_unwanted_columns(df)
                # 5) add these AFTER trim, so they aren't dropped
                df["Brand"] = brand
                df["Source"] = source_name

                all_rows.append(df)


            elif isinstance(match, list) and match:

                df = pd.DataFrame(match)
                df = _add_price_columns(df, source_name, rates)

                if source_name == "legemidler":
                    df = _postprocess_legemidler(df)

                brand_first = str(brand).strip().lower().split()[0] if isinstance(brand, str) else ""
                inp_strength = strength_lookup.get(brand_first, "")

                df = _keep_best_strength_rows(df, inp_strength, source_name)
                df = filter_columns_by_source(df, source_name)
                df = _drop_unwanted_columns(df)

                df["Brand"] = brand
                df["Source"] = source_name

                all_rows.append(df)

            elif isinstance(match, dict):

                available = match.get("available", None)
                submatches = match.get("matches", [])

                if (available is False or available is None) and not submatches:
                    continue

                if isinstance(submatches, list) and submatches:

                    df = pd.DataFrame(submatches)
                    df = _add_price_columns(df, source_name, rates)

                    if source_name == "legemidler":
                        df = _postprocess_legemidler(df)

                    brand_first = str(brand).strip().lower().split()[0] if isinstance(brand, str) else ""
                    inp_strength = strength_lookup.get(brand_first, "")

                    df = _keep_best_strength_rows(df, inp_strength, source_name)
                    df = filter_columns_by_source(df, source_name)
                    df = _drop_unwanted_columns(df)

                    df["Brand"] = brand
                    df["Source"] = source_name
                    all_rows.append(df)

                else:
                    summary_info = {
                        "Brand": brand,
                        "Source": source_name,
                        "available": available,
                        "reason": match.get("reason", "No details")
                    }
                    all_rows.append(pd.DataFrame([summary_info]))

    if not all_rows:
        print("No matches found — skipping sheet write.")
        return

    final_df = pd.concat(all_rows, ignore_index=True)
    # Semantic harmonization with numeric-aware conflict detection
    final_df, groups_with_conflict, conflict_cols = harmonize_semantic_columns_numeric_aware(final_df)

    # Collapse to one canonical column per semantic group + rename Brand
    final_df = collapse_semantic_variants(final_df)
    # Overwrite Brand with the value from input rows (by Ref or substring fallback)
    final_df = normalize_brand_using_input(final_df, rows)

    # Force display brand to the first word of the (mapped) input brand
    if "Brand" in final_df.columns:
        final_df["Brand"] = final_df["Brand"].apply(
            lambda s: s.split()[0] if isinstance(s, str) and s.strip() else s
        )

    # --- Attach requested qty from get_input_rows() ---
    if "Brand" in final_df.columns:
        _final_bfirst = final_df["Brand"].astype(str).str.strip().str.split().str[0].str.lower()
    else:
        _final_bfirst = pd.Series("", index=final_df.index, dtype="object")

    if "Strength" in final_df.columns:
        _final_snorm = final_df["Strength"].astype(str).map(_norm_strength_text)
    else:
        _final_snorm = pd.Series("", index=final_df.index, dtype="object")

    _final_key = _final_bfirst + "||" + _final_snorm

    # try composite key first, then fall back to brand-only
    final_df["Quantatity requested"] = (
        _final_key.map(qty_map_by_key)
        .combine_first(_final_bfirst.map(qty_map_by_brand))
    )

    # --- Sort: Brand → Source → Price EUR (desc). Preserve first-seen order. ---
    price_col  = "Price EUR"
    brand_col  = "Brand"  if "Brand"  in final_df.columns else None
    source_col = "Source" if "Source" in final_df.columns else None

    if brand_col and source_col and price_col in final_df.columns:
        # stable brand order by first appearance
        brand_codes, _ = pd.factorize(final_df[brand_col].astype(str), sort=False)
        final_df["_brand_order"] = brand_codes

        # stable source order within each brand
        final_df["_source_order_within_brand"] = (
            final_df.groupby(brand_col, dropna=False, observed=True)[source_col]
                    .transform(lambda s: pd.factorize(s.astype(str), sort=False)[0])
        )

        # be sure price is numeric
        final_df[price_col] = pd.to_numeric(final_df[price_col], errors="coerce")

        tie_breakers = [c for c in ["Pack Size", "Strength", "Country"] if c in final_df.columns]

        final_df = (
            final_df.sort_values(
                by=["_brand_order", "_source_order_within_brand", price_col] + tie_breakers,
                ascending=[ True,            True,                 False     ] + [True]*len(tie_breakers),
                na_position="last"
            )
            .reset_index(drop=True)
        )

        final_df.drop(columns=["_brand_order", "_source_order_within_brand"], inplace=True, errors="ignore")
    else:
        print("Skip sorting — needed cols missing:",
              {"brand": brand_col is not None, "source": source_col is not None, "price": price_col in final_df.columns})


    # --- Reorder columns: put known ones first, then everything else ---
    preferred_cols = [
        "Ref",
        "Source",
        "Supplier",
        "Quote Date",
        "Brand",
        "Active ingredient",
        "Strength",
        "Form",
        "Pack Size",
        "MA Holder",
        "Country",
        "ATC code",
        "Price EUR",
        "Quantatity requested",
        "qty. Sold last 12 months - Retail",
        "Wholesale Purchase price",
        "Maks AIP Gyldig",
        "Data received",
    ]

    cols_present = [c for c in preferred_cols if c in final_df.columns]
    other_cols = [c for c in final_df.columns if c not in cols_present]
    final_df = final_df[cols_present + other_cols]

    # --- Ensure every input brand appears; add "no match" placeholders ---
    # Make sure Source exists so we can fill it in placeholders
    if "Source" not in final_df.columns:
        final_df["Source"] = pd.NA

    # First-word brand from input (keep original casing from rows)
    rows_first = (
        rows["Brand name"].astype(str).str.strip().str.split().str[0]
    )

    # Existing brands in the result (compare case-insensitively)
    existing_first_lower = (
        final_df["Brand"].astype(str).str.strip().str.split().str[0].str.lower()
    )
    missing_mask = ~rows_first.str.lower().isin(existing_first_lower)
    missing_firsts = rows_first[missing_mask].dropna().unique().tolist()

    placeholders = []
    for b in missing_firsts:
        ph = _blank_row(final_df, key_col="Brand")
        ph.at[ph.index[0], "Brand"] = b  # show the brand name
        ph.at[ph.index[0], "Source"] = "no match"  # mark no match
        placeholders.append(ph)

    if placeholders:
        final_df = pd.concat([final_df] + placeholders, ignore_index=True)

    # --- Add blank rows only between brands ---
    spaced_df = final_df
    if add_spaces and "Brand" in final_df.columns:
        groups = list(final_df.groupby("Brand", dropna=False, observed=True, sort=False))
        parts = []
        for i, (brand, grp) in enumerate(groups):
            if grp.dropna(how="all").empty:
                # if somehow empty, still show a header-like row
                parts.append(_blank_row(final_df, key_col="Brand").assign(Brand=brand, Source="no match"))
            else:
                parts.append(grp)

            # separator between brands (not after the last one)
            if i < len(groups) - 1:
                parts.append(_blank_row(final_df, key_col="Brand"))

        spaced_df = pd.concat(parts, ignore_index=True)

    # write spaced_df instead of final_df
    write_mode = "a" if os.path.exists(output_path) else "w"
    with pd.ExcelWriter(output_path, engine="openpyxl", mode=write_mode) as writer:
        spaced_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    print(f"Wrote {len(spaced_df)} rows to sheet '{sheet_name[:31]}' in: {output_path}")

def _blank_row(df: pd.DataFrame, key_col: str | None = None) -> pd.DataFrame:
    """
    Create a 1-row separator with dtype-appropriate missing values.
    To avoid the FutureWarning about all-NA entries in concat, set key_col to "".
    """
    data = {}
    for col, dtype in df.dtypes.items():
        if key_col and col == key_col:
            # Make the row NOT all-NA (prevents future concat warning)
            data[col] = ""
            continue

        if ptypes.is_float_dtype(dtype):
            data[col] = np.nan
        elif ptypes.is_datetime64_any_dtype(dtype) or ptypes.is_timedelta64_dtype(dtype):
            data[col] = pd.NaT
        elif ptypes.is_integer_dtype(dtype):
            # Nullable integers (Int64 etc.) -> pd.NA; plain int64 can't hold NA, but np.nan
            data[col] = pd.NA if ptypes.is_extension_array_dtype(dtype) else np.nan
        elif ptypes.is_bool_dtype(dtype):
            data[col] = pd.NA if str(dtype) == "boolean" else np.nan
        else:
            # object/string/categorical -> pd.NA is fine
            data[col] = pd.NA

    return pd.DataFrame([data], columns=df.columns)

if __name__ == "__main__":
    from reader import cached_read_excel, get_input_rows

    data = cached_read_excel('Data')
    rows = get_input_rows()

    sourcing_dict = sourcing(rows, data[0])
    product_catalog_dict = product_catalog(rows, data[4])
    legemidler_dict = legemidler(rows, data[1])
    topp_dict = check_availability_no(rows, data[2])
    special_dict = special_access(rows, data[3])

    all_matches = {
        "product_catalog": product_catalog_dict,
        "legemidler": legemidler_dict,
        "special_access": special_dict,
        "sourcing": sourcing_dict,
        "topp 500": topp_dict
    }

    export_matches_to_excel(all_matches)
